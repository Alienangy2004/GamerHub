import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 1. Definir los datos de la tabla de enrutamiento de GamerHub
data = [
    {
        "Red Destino": "0.0.0.0 (Por defecto)",
        "Máscara de Red": "0.0.0.0",
        "Siguiente Salto (Next Hop)": "192.168.50.1",
        "Interfaz de Salida": "FastEthernet 0/0",
        "Tipo de Enrutamiento": "Estático / Default Gateway",
        "Descripción / Propósito": "Todo tráfico hacia Internet pasa obligatoriamente por la inspección del Firewall."
    },
    {
        "Red Destino": "192.168.10.0",
        "Máscara de Red": "255.255.255.0",
        "Siguiente Salto (Next Hop)": "192.168.10.1",
        "Interfaz de Salida": "VLAN 10",
        "Tipo de Enrutamiento": "Directamente conectada",
        "Descripción / Propósito": "Red del departamento de TI e Infraestructura (Equipo-Red)."
    },
    {
        "Red Destino": "192.168.20.0",
        "Máscara de Red": "255.255.255.0",
        "Siguiente Salto (Next Hop)": "192.168.20.1",
        "Interfaz de Salida": "VLAN 20",
        "Tipo de Enrutamiento": "Directamente conectada",
        "Descripción / Propósito": "Red de Sistemas, Soporte y Gestión de SO (Oficinas-SO)."
    },
    {
        "Red Destino": "192.168.30.0",
        "Máscara de Red": "255.255.255.0",
        "Siguiente Salto (Next Hop)": "192.168.30.1",
        "Interfaz de Salida": "VLAN 30",
        "Tipo de Enrutamiento": "Directamente conectada",
        "Descripción / Propósito": "Red de Operaciones, Almacén y Ensamble (Áreas-Equipos)."
    },
    {
        "Red Destino": "192.168.40.0",
        "Máscara de Red": "255.255.255.0",
        "Siguiente Salto (Next Hop)": "192.168.40.1",
        "Interfaz de Salida": "VLAN 40",
        "Tipo de Enrutamiento": "Directamente conectada",
        "Descripción / Propósito": "Red de Comunicación, Diseño y Marketing (Depas-Comunicación)."
    },
    {
        "Red Destino": "192.168.50.0",
        "Máscara de Red": "255.255.255.0",
        "Siguiente Salto (Next Hop)": "192.168.50.1",
        "Interfaz de Salida": "VLAN 50",
        "Tipo de Enrutamiento": "Directamente conectada",
        "Descripción / Propósito": "Zona DMZ donde residen los Servidores Core (IIS Web, DNS, BD)."
    }
]

# 2. Crear el DataFrame y exportarlo a Excel con formato avanzado
df = pd.DataFrame(data)
filename = "tabla-enrutamiento.xlsx"

with pd.ExcelWriter(filename, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name="Enrutamiento GamerHub")
    
    # 3. Estilizar la hoja de cálculo para la entrega de la práctica
    workbook = writer.book
    worksheet = writer.sheets["Enrutamiento GamerHub"]
    
    # Estilos de diseño corporativo (Azul oscuro para encabezados)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11, bold=False)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Aplicar estilos a las celdas de encabezado
    for col_num, column_title in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Aplicar estilos a las filas de datos
    for row in worksheet.iter_rows(min_row=2, max_row=len(data)+1, min_col=1, max_col=6):
        for cell in row:
            cell.font = data_font
            cell.border = thin_border
            if cell.column in [1, 2, 3, 4]:  # IPs e Interfaces centradas
                cell.alignment = center_alignment
            else:
                cell.alignment = left_alignment
                
    # Autoajustar el tamaño de las columnas según el contenido para que no se corte el texto
    for col in worksheet.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        worksheet.column_dimensions[col_letter].width = max(max_len + 3, 15)

print(f" ¡Éxito! El archivo '{filename}' ha sido generado correctamente en la carpeta de tu proyecto.")